import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

# Load the Excel sheet

workbook = load_workbook('TextingData.xlsx')
# select the active sheet
sheet = workbook.active

driver = webdriver.Chrome()
driver.maximize_window()

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    username = row[0]  # based on the name and location from the Excel sheet
    password = row[1]  # based on the name and location from the Excel sheet
    driver.get("https://www.saucedemo.com/")
    time.sleep(10)
    driver.find_element(By.ID, "user-name").send_keys(username)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.ID, "login-button").click()
    time.sleep(10)
    # log out first user and try the other users after login in
    driver.find_element(By.XPATH, "//button[@id='react-burger-menu-btn']").click()
    time.sleep(10)
    driver.find_element(By.XPATH,"//a[@id='logout_sidebar_link']").click()
    time.sleep(10)

driver.quit()
